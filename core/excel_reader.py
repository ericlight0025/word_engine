from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


SUPPORTED_EXCEL_SUFFIXES = {".xlsx", ".xls"}


@dataclass
class ExcelDataset:
    headers: list[str]
    rows: list[dict[str, str]]


def _stringify(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_excel(path: str | Path) -> ExcelDataset:
    file_path = Path(path)
    if file_path.suffix.lower() not in SUPPORTED_EXCEL_SUFFIXES:
        raise ValueError("無法讀取，請確認格式為 .xlsx / .xls")

    workbook = load_workbook(file_path, data_only=True, read_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise ValueError("Excel 內容為空") from exc

    headers = [_stringify(cell) for cell in header_row]
    if not any(headers):
        raise ValueError("Excel 第一列缺少欄位名稱")

    seen: set[str] = set()
    for header in headers:
        if not header:
            raise ValueError("Excel 第一列包含空白欄位名稱")
        if header in seen:
            raise ValueError(f"Excel 欄位名稱重複：{header}")
        seen.add(header)

    data_rows: list[dict[str, str]] = []
    for row in rows_iter:
        values = [_stringify(cell) for cell in row[: len(headers)]]
        if not any(values):
            continue
        padded = values + [""] * (len(headers) - len(values))
        data_rows.append(dict(zip(headers, padded)))

    if not data_rows:
        raise ValueError("Excel 沒有可用資料列")

    return ExcelDataset(headers=headers, rows=data_rows)
